import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import *
import pandas as pd
from tkinter import messagebox
import os
import time
import pyautogui
from pathlib import Path
from bs4 import BeautifulSoup
import pygetwindow as gw
import pandas.io.clipboard as cb
import sys
import openpyxl
from datetime import datetime, timedelta
import shutil
import win32com.client # 调用wps接口
from openpyxl import load_workbook

class ZA_RPA:
    def __init__(self,bank_num_list_path, output_text, xs_path):
        self.bank_num_list_path = bank_num_list_path
        self.output_text = output_text

        
        df_path = self.bank_num_list_path
        self.xs_path = xs_path
        self.output_html_path = os.path.join(os.path.dirname(df_path), 'html')
        self.output_main_path = os.path.join(os.path.dirname(df_path), 'main')
        if not os.path.exists(self.output_html_path): 
            os.mkdir(self.output_html_path)
        if not os.path.exists(self.output_main_path):  
            os.mkdir(self.output_main_path)


    def chrome_equals(self, wn):
        all_windows = gw.getWindowsWithTitle('Google Chrome')
        chrome_tab = ''

        if len(all_windows) > 1:
            self.output_text.config(state=tk.NORMAL)  # 允许编辑文本框以便插入文本
            self.output_text.insert(tk.END, f"请关闭多余的窗口\n")  # 示例输出
            self.output_text.config(state=tk.DISABLED)  # 禁止编辑，使其只读

        for window in all_windows:
            print(window.title)
            window.maximize()
            window.activate()
            chrome_tab = window.title
        
        if wn in chrome_tab:
            return True
        else:
            return False
        

    def click_action(self, x, y, sleep_time=0, scroll_xp=0):
        x = int(x)
        y = int(y)
        sleep = int(sleep_time)
        scroll_xp = int(scroll_xp)
        pyautogui.click(x, y)
        print(x, y, sleep, scroll_xp)
        time.sleep(sleep)
        pyautogui.scroll(scroll_xp)
        time.sleep(1)


    def download_execl(self, xy, bn):
        self.click_action(xy.loc['下载按钮'][0], xy.loc['下载按钮'][1], xy.loc['下载按钮'][2]) # 点击下载
        time.sleep(1)
        if self.chrome_equals('银行卡结果列表') or self.chrome_equals('第三方明细查询详情'):
            return False
        if not self.chrome_equals('无标题'):
            self.output_text.config(state=tk.NORMAL)  # 允许编辑文本框以便插入文本
            self.output_text.insert(tk.END, f"卡号下载错误中断:{bn}\n")  # 示例输出
            self.output_text.config(state=tk.DISABLED)  # 禁止编辑，使其只读
            return True
        self.click_action(xy.loc['下载确定'][0], xy.loc['下载确定'][1], xy.loc['下载确定'][2]) # 点击下载确定
        return False
    

    def download_html(self, bank_name, save_path):
        path = os.path.join(save_path, f'{bank_name}.txt')
        # print(path)
        if os.path.exists(path):
            return
        else:
            pyautogui.hotkey(['ctrl', 'u'])
            time.sleep(3) # 等待加载页面
            pyautogui.hotkey(['ctrl', 'a'])
            pyautogui.hotkey(['ctrl', 'c'])
            html_txt = cb.paste()

            with open(path, 'w') as f:
                f.write(html_txt)

            pyautogui.hotkey(['ctrl', 'w']) 
            time.sleep(2) # 等待关闭


    # 查询当前卡调单记录
    def retrieve_action(self, xy, bn):  
        self.click_action(xy.loc['卡号文本框'][0], xy.loc['卡号文本框'][1], xy.loc['卡号文本框'][2])
        pyautogui.hotkey(['ctrl', 'a'])
        pyautogui.typewrite(bn)
        self.click_action(xy.loc['搜索按钮'][0], xy.loc['搜索按钮'][1], xy.loc['搜索按钮'][2], xy.loc['搜索按钮'][3])



class Card:  
    def __init__(self, name, card_number, bank, start_time, end_time):  
        self.main_category = '个人' 
        self.name = name  
        self.card_number = card_number  
        self.bank = bank  
        self.start_time = datetime.strptime(start_time.split(' ')[0], '%Y-%m-%d')  
        self.end_time = datetime.strptime(end_time.split(' ')[0], '%Y-%m-%d') 
        self.split_method = '按月拆分'

    def get_main_category(self):  
        return self.main_category  

    def set_main_category(self, main_category):  
        self.main_category = main_category  

    def get_name(self):  
        return self.name  

    def set_name(self, name):  
        self.name = name  

    def get_card_number(self):  
        return self.card_number  

    def set_card_number(self, card_number):  
        self.card_number = card_number  

    def get_bank(self):  
        return self.bank  

    def set_bank(self, bank):  
        self.bank = bank  

    def get_start_time(self):  
        return self.start_time  

    def set_start_time(self, start_time):  
        self.start_time = start_time  

    def get_end_time(self):  
        return self.end_time  

    def set_end_time(self, end_time):  
        self.end_time = end_time  

    def get_split_method(self):  
        return self.split_method  

    def set_split_method(self, split_method):  
        self.split_method = split_method


class App:
    def __init__(self, master):
        self.master = master

        # 创建左侧面板
        self.left_frame = tk.Frame(master, bg="lightblue", width=200)
        self.left_frame.pack(side=tk.LEFT, fill=tk.Y)

        # 创建右侧面板
        self.right_frame = tk.Frame(master, width=400)
        self.right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        # 左侧面板上的按钮
        self.buttons = []
        buttons_name = ['流水下载', '脚本配置', '信息提取', '调单填写', '主体提取']
        for i, bnt_name in zip(range(len(buttons_name)), buttons_name):
            button = ttk.Button(self.left_frame, text=f"{bnt_name}", command=lambda i=i: self.show_panel(i))
            button.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=3)
            self.buttons.append(button)


        global BASE_DIR
        global parameters_df
        BASE_DIR = os.path.dirname(os.path.realpath(sys.argv[0]))
        # BASE_DIR = 'C:\\Users\\18428\\Documents\\代码\\RPA程序\\za_rpa_2.0.0'
        self.file_path = os.path.join(BASE_DIR, 'sources\\参数.xlsx')
        parameters_df = pd.read_excel(self.file_path, dtype=str)
        global table_text
        table_text = [[StringVar(value=l) for l in r] for r in parameters_df.iloc[:, 1:].values.tolist()]

        self.bankcard_list = []
        self.bankcard_list_i = 0
        global rpa

        self.not_modify_bank = pd.DataFrame()

        # 初始化右侧面板的内容 默认为第一个
        self.show_panel(0)

    def show_panel_1(self):
        # 第一个界面：下载程序
        def browse_files():
            """当点击按钮时调用此函数以打开文件对话框并获取xlsx文件路径"""
            filename = filedialog.askopenfilename(initialdir="/", title="Select a File",
                                                filetypes=(("Text files", "*.xlsx"), ("All files", "*.*")))
            # 将选择的文件路径显示在输入框中
            directory_entry.delete(0, tk.END)
            directory_entry.insert(0, filename)

        v1 = StringVar()
        directory_entry = ttk.Entry(self.right_frame, textvariable=v1, width=30)
        directory_entry.grid(row=0, column=0, padx=10, pady=10)
        browse_button = ttk.Button(self.right_frame, text="浏览", command=browse_files)
        browse_button.grid(row=0, column=1, padx=10, pady=10,sticky=tk.W)

        # 第二行第一列：Label
        label2 = tk.Label(self.right_frame, text="：从当前卡号下载(默认空为第一张)")
        label2.grid(row=1, column=1, padx=10, pady=10, sticky=tk.W)

        # 第二行第二列：输入框
        entry2 = tk.Entry(self.right_frame, width=30)
        entry2.insert(0, '')  # 在Entry中插入默认值
        entry2.grid(row=1, column=0, padx=10, pady=10)

        # 第三行第一列：Label
        label3 = tk.Label(self.right_frame, text="：一张卡几条下载记录")
        label3.grid(row=2, column=1, padx=10, pady=10, sticky=tk.W)

        # 第三行第二列：输入框
        entry3 = tk.Entry(self.right_frame, width=30)
        entry3.insert(0, '1')  # 在Entry中插入默认值
        entry3.grid(row=2, column=0, padx=10, pady=10)

        # 输出文本框
        output_text = tk.Text(self.right_frame, wrap=tk.WORD, state=tk.DISABLED, height=15, width=60)
        output_scrollbar = tk.Scrollbar(self.right_frame, command=output_text.yview)
        output_text.configure(yscrollcommand=output_scrollbar.set)
        output_text.grid(row=3, column=0, columnspan=2, padx=10, pady=10, sticky=tk.NSEW)
        output_scrollbar.grid(row=3, column=2, sticky=tk.NS)

        # 初始化追踪状态
        self.is_tracking = False

        def start_tracking():
            if not self.is_tracking:
                # print(self.file_path)
                za_rpa = ZA_RPA(v1.get(), output_text, self.file_path)
                input_df = pd.read_excel(v1.get(), dtype=str)
                if '卡号' not in input_df.columns:
                    output_text.config(state=tk.NORMAL)  # 允许编辑文本框以便插入文本
                    output_text.insert(tk.END, f"输入文件不存在卡号列\n")  # 示例输出
                    output_text.config(state=tk.DISABLED)  # 禁止编辑，使其只读
                else:
                    self.bankcard_list = input_df['卡号'].to_list()
                    self.bankcard_list_i = 0 if entry2.get() == '' else self.bankcard_list.index(entry2.get())
                    self.is_tracking = True
                    
                    b_l = len(self.bankcard_list)
                    # print("一张卡下载时间：", jg)
                    output_text.config(state=tk.NORMAL)  # 允许编辑文本框以便插入文本
                    output_text.insert(tk.END, f"3秒后运行\n")  # 示例输出
                    output_text.config(state=tk.DISABLED)  # 禁止编辑，使其只读
                    self.right_frame.after(3000, lambda : update_coordinates(za_rpa, self.bankcard_list, self.bankcard_list_i, b_l))
                # self.master.iconify()

        def update_coordinates(za_rpa, bankcard_list, bankcard_list_i, b_l):
            # za_rpa = ZA_RPA(v1.get(), output_text, self.file_path)
            if not self.is_tracking:
                output_text.config(state=tk.NORMAL)  # 允许编辑文本框以便插入文本
                output_text.insert(tk.END, f"已停止下载，下一张卡：{bankcard_list[bankcard_list_i]}\n")  # 示例输出
                output_text.config(state=tk.DISABLED)  # 禁止编辑，使其只读
                return

            if bankcard_list_i != b_l:
                rpa = za_rpa
                bank_num = bankcard_list[bankcard_list_i]
                num_f = int(entry3.get())
                xy = pd.read_excel(self.file_path).set_index('index')
                view_s = int(xy.loc['查看按钮间距'][1])
                output_text.config(state=tk.NORMAL)  # 允许编辑文本框以便插入文本
                output_text.insert(tk.END, f"卡号:{bank_num}，正在下载\n")  # 示例输出
                output_text.config(state=tk.DISABLED)  # 禁止编辑，使其只读

                if not rpa.chrome_equals('公安部涉赌资金交易查控平台'):
                    output_text.config(state=tk.NORMAL)  # 允许编辑文本框以便插入文本
                    output_text.insert(tk.END, f"卡号下载错误中断:{bank_num}，查控平台没有打开\n")  # 示例输出
                    output_text.config(state=tk.DISABLED)  # 禁止编辑，使其只读
                    self.is_tracking = False
                    return
                
                rpa.retrieve_action(xy, bank_num)
                pyautogui.hotkey(['f11'])   # 防止页面不显示
                time.sleep(0.5)
                pyautogui.hotkey(['f11'])
                time.sleep(0.5)
                
                if not rpa.chrome_equals('公安部涉赌资金交易查控平台'):
                    output_text.config(state=tk.NORMAL)  # 允许编辑文本框以便插入文本
                    output_text.insert(tk.END, f"卡号下载错误中断:{bank_num}\n")  # 示例输出
                    output_text.config(state=tk.DISABLED)  # 禁止编辑，使其只读
                    self.is_tracking = False
                    return
                
                x_ck = xy.loc['查看按钮'][0]
                y_ck = xy.loc['查看按钮'][1]
                if num_f == 0:
                    rpa.click_action(x_ck, y_ck, xy.loc['查看按钮'][2])
                    
                    if rpa.chrome_equals('持卡主体详情') or rpa.chrome_equals('银行卡结果列表') or rpa.chrome_equals('第三方明细查询详情') or rpa.chrome_equals('第三方账户主体查询详情'):
                        if rpa.chrome_equals('持卡主体详情') or rpa.chrome_equals('第三方账户主体查询详情'):
                            rpa.download_html(bank_num, rpa.output_main_path)
                        else:
                            rpa.download_html(bank_num, rpa.output_html_path)
                        pyautogui.hotkey(['ctrl', 'w']) # 关闭查看页面
                        time.sleep(1)
                    else:
                        output_text.config(state=tk.NORMAL)  # 允许编辑文本框以便插入文本
                        output_text.insert(tk.END, f"卡号信息下载错误:{bank_num}\n")  # 示例输出
                        output_text.config(state=tk.DISABLED)  # 禁止编辑，使其只读
                        if not rpa.chrome_equals('公安部涉赌资金交易查控平台'):
                            self.is_tracking = False
                            return
                else:
                    i = 0
                    while i < num_f:
                        rpa.click_action(x_ck, y_ck, xy.loc['查看按钮'][2], xy.loc['查看按钮'][3])
                        if rpa.chrome_equals('公安部涉赌资金交易查控平台'):
                            pyautogui.scroll(int(xy.loc['查看按钮'][3]) * -1) # 滚动像素
                            output_text.config(state=tk.NORMAL)  # 允许编辑文本框以便插入文本
                            output_text.insert(tk.END, f"仅下载 {i} 条数据\n")  # 示例输出
                            output_text.config(state=tk.DISABLED)  # 禁止编辑，使其只读
                            break
                        if rpa.chrome_equals('持卡主体详情') or rpa.chrome_equals('第三方账户主体查询详情'):
                            rpa.download_html(bank_num, rpa.output_main_path)
                            pyautogui.hotkey(['ctrl', 'w']) # 关闭查看页面
                            time.sleep(1)
                            y_ck += view_s
                            continue
                        if not (rpa.chrome_equals('银行卡结果列表') or rpa.chrome_equals('第三方明细查询详情')):
                            output_text.config(state=tk.NORMAL)  # 允许编辑文本框以便插入文本
                            output_text.insert(tk.END, f"卡号下载错误中断:{bank_num}\n")  # 示例输出
                            output_text.config(state=tk.DISABLED)  # 禁止编辑，使其只读
                            self.is_tracking = False
                            return

                        rpa.download_html(bank_num, rpa.output_html_path)
                        if rpa.download_execl(xy, bank_num):
                            self.is_tracking = False
                        pyautogui.hotkey(['ctrl', 'w']) # 关闭查看页面
                        time.sleep(1)
                        y_ck += view_s
                        i += 1
                pyautogui.scroll(int(xy.loc['搜索按钮'][3]) * -1) # 滚动像素
                output_text.config(state=tk.NORMAL)  # 允许编辑文本框以便插入文本
                output_text.insert(tk.END, f"卡号:{bank_num}，已经下载完成\n")  # 示例输出
                output_text.yview(tk.END)
                output_text.config(state=tk.DISABLED)  # 禁止编辑，使其只读
            else:
                self.is_tracking = False
                output_text.config(state=tk.NORMAL)  # 允许编辑文本框以便插入文本
                output_text.insert(tk.END, f"{b_l}张卡（账号）已下载完成\n")  # 示例输出
                output_text.yview(tk.END)
                output_text.config(state=tk.DISABLED)  # 禁止编辑，使其只读
                return

            if self.is_tracking:
                self.right_frame.after(2000, lambda : update_coordinates(rpa, bankcard_list, bankcard_list_i+1, b_l))
            else:
                output_text.config(state=tk.NORMAL)  # 允许编辑文本框以便插入文本
                output_text.insert(tk.END, f"已停止下载\n")  # 示例输出
                output_text.config(state=tk.DISABLED)  # 禁止编辑，使其只读


        def stop_tracking():
            # 创建顶级窗口
            top = tk.Toplevel(root)
            screen_width = top.winfo_screenwidth()
            screen_height = top.winfo_screenheight()
            # 窗口大小
            window_width = 500
            window_height = 80
            x_cordinate = int((screen_width/2) - (window_width/2))
            y_cordinate = int((screen_height/2) - (window_height/2))
            top.geometry(f"{window_width}x{window_height}+{x_cordinate}+{y_cordinate}")
            
            # 创建消息标签
            message_label = tk.Label(top, text="已暂停下载")
            message_label.pack(pady=20, padx=20)
            
            # 定时关闭窗口
            top.after(2000, top.destroy)
            self.is_tracking = False

        # 绑定空格键按下事件
        self.master.bind('<F2>', lambda event: stop_tracking())

        # 第四行跨列的按钮
        submit_button = tk.Button(self.right_frame, text="开始运行", command=lambda : start_tracking())
        submit_button.grid(row=4, column=0, padx=10, pady=10)
        submit_button = tk.Button(self.right_frame, text="停止运行", command=lambda : stop_tracking())
        submit_button.grid(row=4, column=1, padx=10, pady=10)


    def show_panel_2(self):
        # 初始化追踪状态
        self.is_tracking = False

        def start_tracking(i):
            if not self.is_tracking:
                self.is_tracking = True
                # print(i)
                update_coordinates(i)

        def update_coordinates(i):
            # print(self.is_tracking)
            x, y = self.right_frame.winfo_pointerxy()
            table_text[i][0].set(str(x))
            table_text[i][1].set(str(y))
            if self.is_tracking:
                self.right_frame.after(100, lambda i=i: update_coordinates(i))  # 每100毫秒更新一次

        def stop_tracking():
            self.is_tracking = False

        # 绑定空格键按下事件
        self.master.bind('<space>', lambda event: stop_tracking())

        # 创建标签：Label
        first_rows = ['参数', 'x', 'y', '等待时间', '滑动距离']
        first_rows_lable = [tk.Label(self.right_frame, text=n) for n in first_rows]
        for i, lable in zip(range(5), first_rows_lable):
            lable.grid(row=0, column=i, padx=10, pady=10, sticky=tk.W)
        
        first_columns = ['卡输入框', '搜索按钮', '查看按钮', '下载按钮', '下载确定']
        first_columns_lable = [tk.Label(self.right_frame, text=n) for n in first_columns]
        for i, lable in zip(range(1,6), first_columns_lable):
            lable.grid(row=i, column=0, padx=10, pady=10, sticky=tk.W)


        # 创建输入框并设置初始值
        self.xy_entry_list = []
        for i in range(5): 
            for j in range(4):
                tk.Entry(self.right_frame, 
                         width=10, 
                         textvariable=table_text[i][j]).grid(row=i+1, column=j+1, padx=10, pady=10)
            # 创建开始追踪按钮
            self.track_button = tk.Button(self.right_frame, text="捕获位置", command=lambda i=i: start_tracking(i))
            self.track_button.grid(row=i+1, column=5, padx=10, pady=10)
        
        def slip_xp_fun(xp):
            messagebox.showinfo('通知窗口', '点击确定后将在3秒后滑动，请及时切换到对应窗口。')
            self.master.iconify()
            time.sleep(3)
            pyautogui.scroll(int(xp.get())) # 向下滚动xxx像素

        
        self.slip_test_lable = tk.Label(self.right_frame, text="测试滑动")
        xp_scroll = StringVar()
        self.slip_test_entry = tk.Entry(self.right_frame, width=10, textvariable=xp_scroll)
        self.slip_test_button = tk.Button(self.right_frame, text="开始滑动", command=lambda xp=xp_scroll:slip_xp_fun(xp))
        self.slip_test_lable.grid(row=6, column=0, padx=10, pady=10)
        self.slip_test_entry.grid(row=6, column=1, padx=10, pady=10)
        self.slip_test_button.grid(row=6, column=2, padx=10, pady=10)
        self.slip_test_lable2 = tk.Label(self.right_frame, text="点击按钮后3秒开始滑动")
        self.slip_test_lable2.grid(row=6, column=3, columnspan=2,padx=10, pady=10)  

        self.view_spaced_lable = tk.Label(self.right_frame, text="查看间距")
        self.view_spaced_entry = tk.Entry(self.right_frame, width=10, textvariable=table_text[5][1])
        self.view_spaced_lable.grid(row=7, column=0, padx=10, pady=10)
        self.view_spaced_entry.grid(row=7, column=1, padx=10, pady=10)

        def save_parameter_files():
            for i in range(5):
                for j in range(4):
                    parameters_df.iloc[i, j+1] = table_text[i][j].get()
            parameters_df.iloc[5, 2] = table_text[5][1].get()
            # parameters_df.to_excel(self.file_path, sheet_name='坐标', index=False)
            # 指定要保存的Excel文件名
            with pd.ExcelWriter(self.file_path, engine='openpyxl', mode='w') as writer:  # 注意这里使用'w'模式覆盖已有文件，或'a'追加到现有文件
                # 将df1保存为名为'Sheet1'的Sheet
                parameters_df.to_excel(writer, sheet_name='坐标', index=False)
            messagebox.showinfo('通知窗口', '配置文件已保存！')

        self.save_button = tk.Button(self.right_frame, text="保存配置", command=save_parameter_files)
        self.save_button.grid(row=8, column=2, columnspan=2,padx=10, pady=10) 


    def show_panel_3(self):
        # 第三个界面：html解析程序
        def browse_files(entry_dir):
            """当点击按钮时调用此函数以打开文件对话框并获取路径"""
            filename = filedialog.askdirectory()
            entry_dir.delete(0, tk.END)
            entry_dir.insert(0, filename)
            
        tk.Label(self.right_frame, text="输入目录").grid(row=0, column=0, padx=10, pady=5,sticky=tk.W)
        directory_entry1 = ttk.Entry(self.right_frame, width=50)
        directory_entry1.grid(row=1, column=0, padx=10, pady=10)
        browse_button1 = ttk.Button(self.right_frame, text="浏览", command=lambda :browse_files(directory_entry1))
        browse_button1.grid(row=1, column=1, padx=10, pady=10,sticky=tk.W)
        tk.Label(self.right_frame, text="保存目录").grid(row=2, column=0, padx=10, pady=5,sticky=tk.W)
        directory_entry2 = ttk.Entry(self.right_frame, width=50)
        directory_entry2.grid(row=3, column=0, padx=10, pady=10)
        browse_button2 = ttk.Button(self.right_frame, text="浏览", command=lambda :browse_files(directory_entry2))
        browse_button2.grid(row=3, column=1, padx=10, pady=10,sticky=tk.W)

        def get_file_path(dir_path, suffix, recursive=True):
        # 如果recursive是True，任意层级。False,选择当前层级。
            dir = Path(dir_path).resolve()
            pattern = f'**/*.{suffix}' if recursive else f'*.{suffix}'
            return [str(p.resolve()) for p in dir.glob(pattern)]
        
        def get_df(html_path):
            path = html_path
            # print(path)

            with open(path, "rb") as f:
                html_content = f.read()
                html_content = html_content.decode('gbk')

            try:
                soup = BeautifulSoup(html_content, 'html.parser')

                table = soup.find('table', {'class': 'detail_table'})
                rows = table.find_all('tr') 
                user_name_bank = rows[18].find_all('td')[1].get_text().strip()
                user_name = rows[2].find_all('td')[1].get_text().strip()
                account_num = rows[18].find_all('td')[3].get_text().strip()
                bank_num = rows[2].find_all('td')[3].get_text().strip()
                bank = rows[1].find_all('td')[3].get_text().strip()
                status = rows[-4].find_all('td')[3].get_text().strip()
                balance =  rows[-3].find_all('td')[1].get_text().strip()
                last_time = rows[-2].find_all('td')[3].get_text().strip()
                id_card_info = rows[22].find_all('td')[1].get_text().strip()
                id_card = rows[22].find_all('td')[3].get_text().strip()
                phone = rows[23].find_all('td')[3].get_text().strip()
                home_address = rows[25].find_all('td')[3].get_text().strip()
                work_address = rows[26].find_all('td')[3].get_text().strip()
                kai_hu_wang_dian = rows[19].find_all('td')[1].get_text().strip()

                if bank_num != Path(path).stem:
                    self.output_text.config(state=tk.NORMAL)  # 允许编辑文本框以便插入文本
                    self.output_text.insert(tk.END, f"提取错误:{path}\n")  # 示例输出
                    self.output_text.config(state=tk.DISABLED)  # 禁止编辑，使其只读
                    rows_data = []
                else:
                    rows_data = [[user_name_bank, user_name, account_num, bank_num, bank, kai_hu_wang_dian, id_card_info, id_card, phone, home_address, work_address, status, balance, last_time, path]]
            except:
                rows_data = []
                self.output_text.config(state=tk.NORMAL)  # 允许编辑文本框以便插入文本
                self.output_text.insert(tk.END, f"提取错误:{path}\n")  # 示例输出
                self.output_text.config(state=tk.DISABLED)  # 禁止编辑，使其只读
            finally:
                df = pd.DataFrame(rows_data, columns=['开户姓名', '姓名', '主账号', '卡号', '开户行', '开户网点', '查询证照类型', '查询证照号码', '联系手机', '住宅地址', '工作单位', '状态', '余额', '最后交易时间', '文件路径']) 
                return df
        
        def html_merge(input, save):
            path = input.get()
            save_dir = save.get()
            if not path.strip() or not save_dir.strip():  # 检查去除前后空格后是否为空
                messagebox.showwarning("警告", "输入框不能为空，请输入内容！")
            else:
                df_list = []
                for html_path in get_file_path(path, 'txt'):
                    # print(html_path)
                    df = get_df(html_path)
                    df_list.append(df)
                df = pd.concat(df_list).reset_index(drop=True)
                df = df[df['卡号'] != '']
                df['余额'] = df['余额'].apply(lambda x: '0' if x == '' or x == '-' else x)
                df['余额'] = df['余额'].astype('float64')
                df.to_excel(f'{save_dir}\\html.xlsx', index=False)
                # df.to_csv(f'{save_dir}\\html.csv', index=False)
                self.output_text.config(state=tk.NORMAL)  # 允许编辑文本框以便插入文本
                self.output_text.insert(tk.END, f"程序结束！！！\n")  # 示例输出
                self.output_text.insert(tk.END, f"结果已保存至html.xlsx\n")  # 示例输出
                self.output_text.config(state=tk.DISABLED)  # 禁止编辑，使其只读


        # 输出文本框
        self.output_text = tk.Text(self.right_frame, wrap=tk.WORD, state=tk.DISABLED, height=10, width=60)
        output_scrollbar = tk.Scrollbar(self.right_frame, command=self.output_text.yview)
        self.output_text.configure(yscrollcommand=output_scrollbar.set)
        self.output_text.grid(row=4, column=0, columnspan=2, padx=10, pady=10, sticky=tk.NSEW)
        output_scrollbar.grid(row=4, column=2, sticky=tk.NS)

        # 第四行跨列的按钮
        submit_button = ttk.Button(self.right_frame, text="开始运行", command=lambda : html_merge(directory_entry1, directory_entry2))
        submit_button.grid(row=5, column=0, columnspan=2, padx=10, pady=10)

    def show_panel_4(self):
        def get_workfile(path):
            workbook_path = path
            workbook = openpyxl.load_workbook(workbook_path)
            sheet = workbook['银行卡明细查询模板']
            return workbook, sheet

        def date_split(start_time, end_time):
            time_periods = []
            current_time = start_time
            while current_time < end_time:
                next_time = current_time + timedelta(days=85)
                if next_time > end_time:
                    next_time = end_time
                time_periods.append((current_time, next_time))
                current_time = next_time
            return time_periods

        def wirite_paperwork(card_list, template_path, output_dir):
            file_index = 1
            row_index = 3

            workbook, sheet = get_workfile(template_path)

            for card in card_list:
                for date in date_split(card.get_start_time(), card.get_end_time()):        
                    sheet.cell(row=row_index, column=1, value=card.get_main_category())
                    sheet.cell(row=row_index, column=2, value=card.get_name())
                    sheet.cell(row=row_index, column=3, value=card.get_card_number())
                    sheet.cell(row=row_index, column=4, value=card.get_bank())
                    sheet.cell(row=row_index, column=5, value=date[0].strftime("%Y-%m-%d"))
                    sheet.cell(row=row_index, column=6, value=date[1].strftime("%Y-%m-%d"))
                    sheet.cell(row=row_index, column=7, value=card.get_split_method())
                    row_index += 1

                    if row_index > 95:
                        workbook.save(os.path.join(output_dir, f'zn_{file_index}.xlsx'))
                        file_index += 1
                        workbook.close()
                        workbook, sheet = get_workfile(template_path)
                        row_index = 3

            workbook.save(os.path.join(output_dir, f'zn_{file_index}.xlsx'))
            workbook.close()

        def del_file(path):
            file_path = path
            if os.path.exists(file_path):
                try:
                    # 尝试删除文件
                    os.remove(file_path)
                    print(f"文件 {file_path} 已被删除。")
                except FileNotFoundError:
                    # 如果文件不存在（可能已经被其他进程删除），忽略错误
                    print(f"文件 {file_path} 不存在，无需删除。")
                except Exception as e:
                    # 处理其他可能的异常
                    print(f"删除文件时发生错误: {e}")

        def run_matching(bank_path):  
            bank_path = bank_path

            SOURCES_DIR = os.path.join(BASE_DIR, 'sources')

            bank_dict_path = os.path.join(SOURCES_DIR, '发卡机构字典.xlsx')
            template_path = os.path.join(SOURCES_DIR, 'zn_template.xlsx')   # 治安调单表模板
            output_dir = os.path.join(BASE_DIR, 'output')
            tmp_dir = os.path.join(BASE_DIR, 'tmp')
            del_file(os.path.join(BASE_DIR, '调单卡_未匹配到银行.xlsx'))
            del_file(os.path.join(BASE_DIR, '调单卡_未匹配到用户.xlsx'))

            bank_df = pd.read_excel(bank_path, dtype=str)
            is_return = False
            for col in ['姓名', '卡号', '发卡机构', '开始时间', '结束时间']:
                if col not in bank_df.columns:
                    output_text.config(state=tk.NORMAL)  # 允许编辑文本框以便插入文本
                    output_text.insert(tk.END, f"《{col}》列不存在\n")  # 输出
                    output_text.config(state=tk.DISABLED)  # 禁止编辑，使其只读
                    is_return = True
            if is_return:
                return
                       
            bank_dict = pd.read_excel(bank_dict_path, dtype=str)
            merged_df = pd.merge(bank_df, bank_dict, how='left', on='发卡机构')
            merged_df['治安发卡机构'].fillna('-', inplace=True)

            t_df = merged_df[['姓名', '卡号', '发卡机构', '治安发卡机构']]
            merged_df = merged_df[merged_df['治安发卡机构'] != '-']

            t_df = t_df[t_df['治安发卡机构'] == '-']
            output_text.config(state=tk.NORMAL)  # 允许编辑文本框以便插入文本
            if t_df.empty:  
                output_text.insert(tk.END, "开户行已全部匹配\n")  # 示例输出
            else:  
                t_df_bank = t_df.drop_duplicates(subset=['发卡机构'])
                t_df_bank.to_excel(os.path.join(BASE_DIR, '调单卡_未匹配到银行.xlsx'), index=False)
                t_df.to_excel(os.path.join(BASE_DIR, '调单卡_未匹配到用户.xlsx'), index=False)
                self.not_modify_bank = t_df_bank
                self.not_modify_bank = self.not_modify_bank[['卡号', '发卡机构', '治安发卡机构']]
                self.not_modify_bank = self.not_modify_bank.rename(columns={'卡号': '银行卡号'})
                output_text.insert(tk.END, f"未匹配银行：共 {t_df_bank.shape[0]} 个银行未匹配\n未匹配银行卡已输出至：{os.path.join(BASE_DIR, '调单卡_未匹配到银行.xlsx')}\n")  # 示例输出
                output_text.insert(tk.END, f"未匹配用户：共 {t_df.shape[0]} 个用户未匹配\n未匹配用户已输出至：{os.path.join(BASE_DIR, '调单卡_未匹配到用户.xlsx')}\n")  # 示例输出
            output_text.config(state=tk.DISABLED)  # 禁止编辑，使其只读

            merged_df = merged_df[['姓名', '卡号', '治安发卡机构', '开始时间', '结束时间']].values
            card_list = [Card(bl[0], bl[1], bl[2], bl[3], bl[4]) for bl in merged_df]

            if not os.path.exists(tmp_dir):  
                    os.mkdir(tmp_dir)  
            else:
                shutil.rmtree(tmp_dir)
                os.mkdir(tmp_dir)
                
            if not os.path.exists(output_dir):  
                    os.mkdir(output_dir)  
            else:
                shutil.rmtree(output_dir)
                os.mkdir(output_dir)

            wirite_paperwork(card_list, template_path, tmp_dir)

            xlsx_path = os.listdir(tmp_dir)
            for p in xlsx_path:
                path = os.path.join(tmp_dir, p)
                application = win32com.client.Dispatch('Excel.Application')
                workbook = application.Workbooks.Open(path)
                worksheet = workbook.Worksheets('银行卡明细查询模板')
                dirname, filename = os.path.split(path)
                worksheet.SaveAs(os.path.join(output_dir, filename))
                workbook.Close()
                application.Quit()
            shutil.rmtree(tmp_dir)
            output_text.config(state=tk.NORMAL)  # 允许编辑文本框以便插入文本
            output_text.insert(tk.END, f"输出文件夹：{output_dir}\n\n\n")  # 示例输出
            output_text.config(state=tk.DISABLED)  # 禁止编辑，使其只读
        
                # 第一个界面：下载程序
        def browse_files():
            """当点击按钮时调用此函数以打开文件对话框并获取xlsx文件路径"""
            filename = filedialog.askopenfilename(initialdir="/", title="Select a File",
                                                filetypes=(("Text files", "*.xlsx"), ("All files", "*.*")))
            # 将选择的文件路径显示在输入框中
            directory_entry.delete(0, tk.END)
            directory_entry.insert(0, filename)

        def modify_dict():
            file_path = os.path.join(BASE_DIR, 'sources', '发卡机构字典.xlsx')
            dict_df = pd.read_excel(file_path, dtype=str)
            new_dict_df = pd.concat([self.not_modify_bank, dict_df])
            try:
                new_dict_df.to_excel(file_path, index=False)

                # 然后加载这个Excel文件，并调整列宽
                workbook = load_workbook(file_path)
                worksheet = workbook.active
                # 调整列宽
                worksheet.column_dimensions['A'].width = 25
                worksheet.column_dimensions['B'].width = 35
                worksheet.column_dimensions['C'].width = 35
                workbook.save(file_path)
            except PermissionError as e:
                messagebox.showinfo('通知窗口', '配置文件已打开！')
                print('文件已打开')
            try:
                os.startfile(file_path)
                print("文件已用默认应用打开。")
            except OSError as e:
                print(f"无法打开文件: {e}")
            self.not_modify_bank = pd.DataFrame()

        # 第二行第一列：Label
        label1 = tk.Label(self.right_frame, text="输入模板文件：")
        label1.grid(row=0, column=0, padx=10, pady=10, sticky=tk.W)

        bp = StringVar()
        directory_entry = ttk.Entry(self.right_frame, textvariable=bp, width=30)
        directory_entry.grid(row=0, column=1, padx=10, pady=10)
        browse_button = ttk.Button(self.right_frame, text="浏览", command=browse_files)
        browse_button.grid(row=0, column=2, padx=10, pady=10,sticky=tk.W)

        submit_button = ttk.Button(self.right_frame, text="开始运行", command=lambda: run_matching(bp.get()))
        submit_button.grid(row=2, column=0, padx=10, pady=10)
        submit_button = ttk.Button(self.right_frame, text="更新字典", command=modify_dict)
        submit_button.grid(row=2, column=1, padx=10, pady=10)

        # 输出文本框
        output_text = tk.Text(self.right_frame, wrap=tk.WORD, state=tk.DISABLED, height=15, width=65)
        output_scrollbar = tk.Scrollbar(self.right_frame, command=output_text.yview)
        output_text.configure(yscrollcommand=output_scrollbar.set)
        output_text.grid(row=1, column=0, columnspan=3, padx=10, pady=10, sticky=tk.NSEW)
        output_scrollbar.grid(row=1, column=4, sticky=tk.NS)

    
    def show_panel_5(self):
        # 第三个界面：html解析程序
        def browse_files(entry_dir):
            """当点击按钮时调用此函数以打开文件对话框并获取路径"""
            filename = filedialog.askdirectory()
            entry_dir.delete(0, tk.END)
            entry_dir.insert(0, filename)
            
        tk.Label(self.right_frame, text="输入目录").grid(row=0, column=0, padx=10, pady=5,sticky=tk.W)
        directory_entry1 = ttk.Entry(self.right_frame, width=50)
        directory_entry1.grid(row=1, column=0, padx=10, pady=10)
        browse_button1 = ttk.Button(self.right_frame, text="浏览", command=lambda :browse_files(directory_entry1))
        browse_button1.grid(row=1, column=1, padx=10, pady=10,sticky=tk.W)
        tk.Label(self.right_frame, text="保存目录").grid(row=2, column=0, padx=10, pady=5,sticky=tk.W)
        directory_entry2 = ttk.Entry(self.right_frame, width=50)
        directory_entry2.grid(row=3, column=0, padx=10, pady=10)
        browse_button2 = ttk.Button(self.right_frame, text="浏览", command=lambda :browse_files(directory_entry2))
        browse_button2.grid(row=3, column=1, padx=10, pady=10,sticky=tk.W)

        def get_file_path(dir_path, suffix, recursive=True):
        # 如果recursive是True，任意层级。False,选择当前层级。
            dir = Path(dir_path).resolve()
            pattern = f'**/*.{suffix}' if recursive else f'*.{suffix}'
            return [str(p.resolve()) for p in dir.glob(pattern)]
        
        def get_wechat_df(html_path):
            path = html_path
    
            with open(path, "rb") as f:
                html_content = f.read()
                html_content = html_content.decode('gbk')

            soup = BeautifulSoup(html_content, 'html.parser')
            df = pd.DataFrame()
            try:
                table = soup.find('table', {'class': 'main_table'})
                colume = [html.get_text().strip() for html in table.find('tr', {'class': 'tit'}).find_all('td', {'class': 'bgc2'})]
                data_all = table.find_all('tr')[1:]
                data_list = []
                for data in data_all:
                    data_info = []
                    for i in data.find_all('td'):
                        data_info.append(i.get_text().strip())
                    data_list.append(data_info)
                df = pd.DataFrame(columns=colume, data=data_list)
            except:
                self.output_text.config(state=tk.NORMAL)  # 允许编辑文本框以便插入文本
                self.output_text.insert(tk.END, f"提取错误:{path}\n")  # 示例输出
                self.output_text.config(state=tk.DISABLED)  # 禁止编辑，使其只读
            finally:
                return df
        
        def html_merge(input, save):
            path = input.get()
            save_dir = save.get()
            if not path.strip() or not save_dir.strip():  # 检查去除前后空格后是否为空
                messagebox.showwarning("警告", "输入框不能为空，请输入内容！")
            else:
                df_list = []
                for html_path in get_file_path(path, 'txt'):
                    # print(html_path)
                    df = get_wechat_df(html_path)
                    df_list.append(df)
                df = pd.concat(df_list).reset_index(drop=True)
                df.to_excel(f'{save_dir}\\main.xlsx', index=False)
                self.output_text.config(state=tk.NORMAL)  # 允许编辑文本框以便插入文本
                self.output_text.insert(tk.END, f"程序结束！！！\n")  # 示例输出
                self.output_text.insert(tk.END, f"结果已保存至main.xlsx\n")  # 示例输出
                self.output_text.config(state=tk.DISABLED)  # 禁止编辑，使其只读


        # 输出文本框
        self.output_text = tk.Text(self.right_frame, wrap=tk.WORD, state=tk.DISABLED, height=10, width=60)
        output_scrollbar = tk.Scrollbar(self.right_frame, command=self.output_text.yview)
        self.output_text.configure(yscrollcommand=output_scrollbar.set)
        self.output_text.grid(row=4, column=0, columnspan=2, padx=10, pady=10, sticky=tk.NSEW)
        output_scrollbar.grid(row=4, column=2, sticky=tk.NS)
        
        # 第四行跨列的按钮
        submit_button = ttk.Button(self.right_frame, text="开始运行", command=lambda : html_merge(directory_entry1, directory_entry2))
        submit_button.grid(row=5, column=0, columnspan=2, padx=10, pady=10)


    def show_panel(self, index):
        """显示对应的右侧面板内容"""
        # 销毁右面板的内容
        for widget in self.right_frame.winfo_children():
            widget.destroy()

        if index == 0:
            self.show_panel_1()
        elif index == 1:
            self.show_panel_2()
        elif index == 2:
            self.show_panel_3()
        elif index == 3:
            self.show_panel_4()
        elif index == 4:
            self.show_panel_5()


if __name__ == "__main__":
    root = tk.Tk()
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    # 窗口大小
    window_width = 800
    window_height = 500
    x_cordinate = int((screen_width/2) - (window_width/2))
    y_cordinate = int((screen_height/2) - (window_height/2))
    root.geometry(f"{window_width}x{window_height}+{x_cordinate}+{y_cordinate}")
    root.title('治安工具箱')
    app = App(root)
    root.mainloop()